from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("/Users/kei/Downloads/Tüpfeltabelle.xlsx")
SHEET_NAME = "RAW"
OUTPUT = Path("src/data.js")

COLOR_RULES = [
    ("white", r"\bwei(?:ß|ss)|weiss|weiß"),
    ("yellow", r"gelb|creme"),
    ("orange", r"orange"),
    ("brown", r"braun|sand"),
    ("red", r"\brot|karmin|ziegel|blut"),
    ("blue", r"blau"),
    ("green", r"gr(?:ü|u)n"),
    ("black", r"schwarz"),
    ("purple", r"violett|lila"),
    ("pink", r"rosa|pink"),
    ("gray", r"grau"),
    ("turquoise", r"t(?:ü|u)rkis"),
]

PRECIPITATE_MARKERS = [
    "niederschlag",
    " ns",
    "trübung",
    "trubung",
    "häutchen",
    "hautchen",
    "glibberig",
    "fein",
    "ül",
    "ü.l",
    " l in ",
    "löslich",
    "loslich",
    "kaum sichtbar",
    "selten",
]

OTHER_MARKERS = [
    "gasentwicklung",
    "geruch",
    "neutralisation",
    "oxidation",
    "entfärbung",
    "entfarbung",
    "puffer",
    "papier",
    "basisch",
    "sauer",
    "neutral",
    "seifig",
    "ringprobe",
    "nachweise",
    "experimentelle",
    "fühlt",
    "fuehlt",
    "riechen",
    "ph ",
    "mehr mno4",
    "minus mal minus",
]

SOLUTION_MARKERS = [
    "lösung",
    "losung",
    "färbung",
    "farbung",
    "komplex",
    "eigenfarbe",
]

PROPERTY_COLUMNS = {"Eigenfarbe", "Flammenfärbung", "pH-Papier", "extra Hinweise"}
PROPERTY_ROWS = {"Eigenfarbe", "pH-Papier", "extra Hinweise", "I2 + H+", "I2"}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return normalized or "blank"


def extract_colors(text: str) -> list[str]:
    lowered = text.lower()
    colors = []
    for color_id, pattern in COLOR_RULES:
        if re.search(pattern, lowered):
            colors.append(color_id)
    return colors


def has_any(text: str, markers: list[str]) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in markers)


def classify(primary: str, secondary: str, observation: str) -> str:
    if not observation:
        return "none"

    lowered = observation.lower()
    if primary in PROPERTY_ROWS or secondary in PROPERTY_COLUMNS:
        if secondary == "Eigenfarbe":
            return "solution"
        return "other"

    if " ns" in f" {lowered}" or "niederschlag" in lowered:
        return "precipitate"
    if has_any(observation, OTHER_MARKERS):
        return "other"
    if has_any(observation, PRECIPITATE_MARKERS):
        return "precipitate"
    if has_any(observation, SOLUTION_MARKERS):
        return "solution"
    if extract_colors(observation):
        return "precipitate"
    return "other"


def build() -> dict[str, object]:
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook[SHEET_NAME]
    headers = [clean_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    reagents = [header for header in headers[1:] if header]
    ions = []
    reactions = []

    for row in range(2, sheet.max_row + 1):
        primary = clean_text(sheet.cell(row, 1).value)
        if not primary:
            continue
        ions.append(primary)

        for col, secondary in enumerate(reagents, start=2):
            observation = clean_text(sheet.cell(row, col).value)
            outcome = classify(primary, secondary, observation)
            reaction_id = f"{slug(primary)}__{slug(secondary)}"
            reactions.append(
                {
                    "id": reaction_id,
                    "primary": primary,
                    "secondary": secondary,
                    "observation": observation,
                    "outcome": outcome,
                    "colors": extract_colors(observation),
                    "sourceCell": f"{SHEET_NAME}!{sheet.cell(row, col).coordinate}",
                    "inCoreDrill": primary not in PROPERTY_ROWS
                    and secondary not in PROPERTY_COLUMNS
                    and primary != secondary,
                }
            )

    return {
        "source": {
            "workbook": WORKBOOK.name,
            "sheet": SHEET_NAME,
            "notes": "Blank cells and whitespace-only cells are normalized as no visible reaction.",
        },
        "ions": ions,
        "reagents": reagents,
        "reactions": reactions,
    }


def write_module(data: dict[str, object]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(data, ensure_ascii=False, indent=2)
    OUTPUT.write_text(
        "export const REACTION_DATA = "
        + serialized
        + ";\n\n"
        + "export const REACTIONS = REACTION_DATA.reactions;\n"
        + "export const IONS = REACTION_DATA.ions;\n"
        + "export const REAGENTS = REACTION_DATA.reagents;\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    write_module(build())
